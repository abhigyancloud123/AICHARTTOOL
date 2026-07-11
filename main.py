import warnings
warnings.filterwarnings("ignore", category=FutureWarning)

import os
import re
import json
import time
import datetime
from io import StringIO
from typing import List, Literal, Dict

import requests
from bs4 import BeautifulSoup
import pandas as pd
import matplotlib
try:
    import tkinter  # noqa: F401
    matplotlib.use("TkAgg", force=True)
except Exception:
    try:
        import PyQt5  # noqa: F401
        matplotlib.use("QtAgg", force=True)
    except Exception:
        matplotlib.use("Agg", force=True)

import matplotlib.pyplot as plt
from google import genai
from pydantic import BaseModel, Field
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
try:
    from openpyxl.chart.legend import Legend
except ImportError:
    Legend = None
try:
    import yt_dlp
    YT_DLP_AVAILABLE = True
except ImportError:
    YT_DLP_AVAILABLE = False

try:
    import tkinter as tk
    from tkinter import ttk, messagebox
except Exception:
    tk = None
    ttk = None
    messagebox = None

script_dir = os.path.dirname(os.path.abspath(__file__))
os.chdir(script_dir)


# ============================================
# CONFIG & SCHEMAS
# ============================================
# API key can be set via environment variable GEMINI_API_KEY or via a local file.
# For security, avoid hardcoding keys in source.
API_KEY = os.environ.get("GEMINI_API_KEY", "")
MODEL_NAME = "gemini-2.5-flash"
DEFAULT_BAR_LINE_LIMIT = 10
DEFAULT_PIE_LIMIT = 6

YOUTUBE_PATTERN = re.compile(r"(youtube\.com/watch|youtu\.be/|youtube\.com/shorts/)")


class ChartAnalysis(BaseModel):
    label_column: str = Field(description="The exact name of the column containing names or categories.")
    value_columns: List[str] = Field(description="A list of exact column names containing numeric values to plot. If tracking trends over time (e.g., years), include all relevant columns.")
    title: str = Field(description="A short, catchy title for the generated chart.")
    x_axis_title: str = Field(description="Clear label for the horizontal axis based on the columns.")
    y_axis_title: str = Field(description="Clear label for the vertical axis detailing what the numbers represent.")
    recommended: Literal["bar", "line", "pie"] = Field(description="Chart type recommendation (line for time-series, pie for parts-of-a-whole under 8 items, bar for discrete item comparisons).")
    show_legend: bool = Field(description="True if there are multiple lines/bars or if clarity requires a legend; otherwise False.")
    reason: str = Field(description="Brief reason for this choice.")


class FlatDataAnalysis(BaseModel):
    labels: List[str] = Field(description="The array of data category labels.")
    values: List[float] = Field(description="The numeric array matching the labels sequence.")
    title: str = Field(description="A concise chart title.")
    x_axis_title: str = Field(description="Clear label for the horizontal axis.")
    y_axis_title: str = Field(description="Clear label for the vertical axis.")
    recommended: Literal["bar", "line", "pie"] = Field(description="The recommended chart type configuration.")
    show_legend: bool = Field(description="True if legend is needed; otherwise False.")
    reason: str = Field(description="Brief reason for the recommendation.")


# ============================================
# SMALL UI HELPERS
# ============================================
_root = None


def _get_root():
    global _root
    if _root is None and tk is not None:
        _root = tk.Tk()
        _root.withdraw()
    return _root


def popup_ask_text(title: str, prompt: str, default: str = "") -> str:
    if tk is None or ttk is None:
        return input(prompt).strip()

    root = _get_root()
    win = tk.Toplevel(root)
    win.title(title)
    win.configure(bg="#f4f5f7")
    win.attributes("-topmost", True)
    win.resizable(False, False)
    win.update_idletasks()
    sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
    w, h = 480, 190
    x, y = (sw - w) // 2, (sh - h) // 2
    win.geometry(f"{w}x{h}+{x}+{y}")

    tk.Label(win, text=prompt, bg="#f4f5f7", fg="#1e1e1e", wraplength=440, justify="left", font=("Segoe UI", 11)).pack(pady=(20, 10), padx=20)
    entry_var = tk.StringVar(value=default)
    entry = tk.Entry(win, textvariable=entry_var, font=("Segoe UI", 11), width=40)
    entry.pack(pady=6, padx=20)
    entry.focus_set()

    result = {"value": None}

    def submit(event=None):
        result["value"] = entry_var.get().strip()
        win.destroy()

    entry.bind("<Return>", submit)
    ttk.Button(win, text="Continue", command=submit).pack(pady=12)
    win.wait_window()
    return result["value"] or ""


def popup_ask_choice(title: str, prompt: str, options: list) -> str:
    if tk is None or ttk is None:
        return input(prompt).strip().lower()

    root = _get_root()
    win = tk.Toplevel(root)
    win.title(title)
    win.configure(bg="#f4f5f7")
    win.attributes("-topmost", True)
    win.resizable(False, False)
    win.update_idletasks()
    sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
    w, h = 480, 180 + max(0, (len(options) - 1) // 2) * 50
    x, y = (sw - w) // 2, (sh - h) // 2
    win.geometry(f"{w}x{h}+{x}+{y}")

    tk.Label(win, text=prompt, bg="#f4f5f7", fg="#1e1e1e", wraplength=440, justify="left", font=("Segoe UI", 11)).pack(pady=(20, 12), padx=20)

    chosen = {"value": None}

    def pick(opt):
        chosen["value"] = opt
        win.destroy()

    frame = tk.Frame(win, bg="#f4f5f7")
    frame.pack()
    for i, opt in enumerate(options):
        row, col = divmod(i, 2)
        ttk.Button(frame, text=opt.capitalize(), width=16, command=lambda o=opt: pick(o)).grid(row=row, column=col, padx=8, pady=6)

    win.wait_window()
    return chosen["value"] or ""


def popup_ask_from_list(title: str, prompt: str, items: list) -> int:
    if tk is None or ttk is None:
        return 0

    root = _get_root()
    win = tk.Toplevel(root)
    win.title(title)
    win.configure(bg="#f4f5f7")
    win.attributes("-topmost", True)
    win.resizable(False, False)
    win.update_idletasks()
    sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
    h = min(520, 180 + len(items) * 58)
    w = 620
    x, y = (sw - w) // 2, (sh - h) // 2
    win.geometry(f"{w}x{h}+{x}+{y}")

    tk.Label(win, text=prompt, bg="#f4f5f7", fg="#1e1e1e", wraplength=580, justify="left", font=("Segoe UI", 11, "bold")).pack(pady=(15, 8), padx=20)

    canvas = tk.Canvas(win, bg="#f4f5f7", highlightthickness=0)
    canvas.pack(fill="both", expand=True, padx=10, pady=(0, 10))
    frame = tk.Frame(canvas, bg="#f4f5f7")
    canvas.create_window((0, 0), window=frame, anchor="nw")

    chosen = {"value": None}

    def pick(idx):
        chosen["value"] = idx
        win.destroy()

    for i, desc in enumerate(items):
        row = tk.Frame(frame, bg="white", relief="solid", borderwidth=1)
        row.pack(fill="x", padx=6, pady=4)
        tk.Label(row, text=desc, bg="white", fg="#1e1e1e", wraplength=470, justify="left", anchor="w", font=("Segoe UI", 10)).pack(side="left", padx=10, pady=8, fill="x", expand=True)
        ttk.Button(row, text="Use", command=lambda idx=i: pick(idx)).pack(side="right", padx=8, pady=6)

    frame.update_idletasks()
    canvas.configure(scrollregion=canvas.bbox("all"))
    win.wait_window()
    return chosen["value"] if chosen["value"] is not None else 0


def popup_info(title: str, message: str):
    if messagebox is not None:
        messagebox.showinfo(title, message)
    else:
        print(message)


def popup_warn(title: str, message: str):
    if messagebox is not None:
        messagebox.showwarning(title, message)
    else:
        print(message)


def step(msg: str):
    print(f"\n> {msg}")


def info(msg: str):
    print(f"  {msg}")


def warn(msg: str):
    print(f"  ! {msg}")


def is_youtube_url(url: str) -> bool:
    return bool(YOUTUBE_PATTERN.search(url))


def choose_item_count(data: dict, chart_type: str) -> int:
    total = len(data["labels"])
    default = DEFAULT_PIE_LIMIT if chart_type in ("pie", "doughnut") else DEFAULT_BAR_LINE_LIMIT
    default = min(default, total)

    choice = popup_ask_text(
        "How many items?",
        f"{total} items were found.\nType a number, type 'all', or leave blank to use the default ({default}).",
        str(default),
    ).strip().lower()

    if choice == "":
        return default
    if choice == "all":
        return total
    try:
        val = int(choice)
        return max(1, min(val, total))
    except ValueError:
        warn(f"Invalid input recognized. Defaulting to top {default} items.")
        return default


def trim_data(data: dict, count: int) -> dict:
    data["labels"] = data["labels"][:count]
    for col in data["series"]:
        data["series"][col] = data["series"][col][:count]
    return data


# ============================================
# DATA VISUALIZATION & EXPORT
# ============================================
def save_chart_image(data: dict, analysis: ChartAnalysis, chart_type: str, filename: str = "chart.png"):
    plt.style.use("seaborn-v0_8-whitegrid")
    fig, ax = plt.subplots(figsize=(14, 8), dpi=140)
    fig.patch.set_facecolor("white")
    ax.set_facecolor("#f8f9fb")

    labels = data["labels"]
    series_dict = data["series"]
    colors = plt.cm.tab10.colors

    if chart_type == "bar":
        x = list(range(len(labels)))
        width = 0.8 / max(1, len(series_dict))
        bars = []
        for i, (series_name, values) in enumerate(series_dict.items()):
            pos = [idx - 0.4 + (i + 0.5) * width for idx in x]
            bar_container = ax.bar(
                pos,
                values,
                width=width,
                label=series_name,
                color=colors[i % len(colors)],
                edgecolor="black",
                linewidth=0.8,
                alpha=0.95,
            )
            bars.append(bar_container)
        ax.set_xticks(x)
        ax.set_xticklabels(labels, rotation=45, ha="right", fontsize=10)
        ax.grid(axis="y", linestyle="--", linewidth=0.7, alpha=0.5)
        if len(labels) <= 15:
            for container in bars:
                for bar in container:
                    height = bar.get_height()
                    if height != 0:
                        ax.text(
                            bar.get_x() + bar.get_width() / 2,
                            height,
                            f"{height:,.0f}",
                            ha="center",
                            va="bottom",
                            fontsize=8,
                            color="black",
                        )

    elif chart_type == "line":
        for i, (series_name, values) in enumerate(series_dict.items()):
            ax.plot(
                labels,
                values,
                marker="o",
                linewidth=2.6,
                markersize=7,
                label=series_name,
                color=colors[i % len(colors)],
            )
        ax.set_xticks(range(len(labels)))
        ax.set_xticklabels(labels, rotation=45, ha="right", fontsize=10)
        ax.grid(True, linestyle="--", linewidth=0.7, alpha=0.5)
        if len(labels) <= 20:
            for line in ax.lines:
                for x_val, y_val in zip(line.get_xdata(), line.get_ydata()):
                    ax.text(x_val, y_val, f"{y_val:,.0f}", ha="center", va="bottom", fontsize=8, color="black")

    elif chart_type == "pie":
        first_col = list(series_dict.keys())[0]
        wedges, texts, autotexts = ax.pie(
            series_dict[first_col],
            labels=labels,
            autopct="%1.1f%%",
            startangle=90,
            wedgeprops={"edgecolor": "white", "linewidth": 1.2},
            pctdistance=0.6,
            textprops={"fontsize": 10},
        )
        for autotext in autotexts:
            autotext.set_color("white")
            autotext.set_fontweight("bold")
        ax.axis("equal")

    ax.set_title(analysis.title, fontsize=16, fontweight="bold", pad=18)
    if chart_type != "pie":
        ax.set_xlabel(analysis.x_axis_title, fontsize=12, labelpad=10)
        ax.set_ylabel(analysis.y_axis_title, fontsize=12, labelpad=10)

    if chart_type != "pie" and (analysis.show_legend or len(series_dict) > 1):
        legend = ax.legend(loc="upper right", frameon=True, shadow=True, facecolor="white", framealpha=0.95)
        legend.get_frame().set_edgecolor("#d0d0d0")

    ax.set_axisbelow(True)
    for spine in ("top", "right"):
        ax.spines[spine].set_visible(False)

    plt.tight_layout()
    plt.savefig(filename, dpi=220, bbox_inches="tight", facecolor=fig.get_facecolor())
    info(f"Chart image saved as {filename}")
    try:
        plt.show(block=True)
    except Exception:
        plt.show()


def save_excel(data: dict, analysis: ChartAnalysis, chart_type: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Insights"

    ws["A1"] = "Label"
    series_names = list(data["series"].keys())

    for col_idx, name in enumerate(series_names, start=2):
        ws.cell(row=1, column=col_idx, value=name)

    for row_idx, label in enumerate(data["labels"], start=2):
        ws.cell(row=row_idx, column=1, value=label)
        for col_idx, name in enumerate(series_names, start=2):
            ws.cell(row=row_idx, column=col_idx, value=data["series"][name][row_idx - 2])

    chart_classes = {"bar": BarChart, "line": LineChart, "pie": PieChart}
    excel_chart = chart_classes.get(chart_type, BarChart)()

    if chart_type == "bar":
        excel_chart.type = "col"
        excel_chart.style = 10
    elif chart_type == "line":
        excel_chart.style = 12

    excel_chart.title = analysis.title
    excel_chart.width = 22
    excel_chart.height = 14

    if chart_type != "pie":
        excel_chart.x_axis.title = analysis.x_axis_title
        excel_chart.y_axis.title = analysis.y_axis_title
        if Legend is not None and (analysis.show_legend or len(series_names) > 1):
            excel_chart.legend = Legend()
            excel_chart.legend.position = "r"

    data_ref = Reference(ws, min_col=2, max_col=len(series_names) + 1, min_row=1, max_row=len(data["labels"]) + 1)
    labels_ref = Reference(ws, min_col=1, min_row=2, max_row=len(data["labels"]) + 1)
    excel_chart.add_data(data_ref, titles_from_data=True)
    excel_chart.set_categories(labels_ref)

    ws.add_chart(excel_chart, "D2")
    ws.column_dimensions["A"].width = 28

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"chart_{timestamp}.xlsx"
    wb.save(filename)
    info(f"Excel file saved as {filename}")
    return filename


def choose_chart_type(recommended_type: str, reason: str) -> str:
    choice = popup_ask_choice(
        "Chart type",
        f"Recommended: {recommended_type.upper()}\nWhy: {reason}\n\nPick a chart type:",
        ["bar", "line", "pie"],
    )
    return choice if choice in ("bar", "line", "pie") else recommended_type


# ============================================
# WEBPAGE PROCESSING WORKFLOW
# ============================================
def fetch_html(url: str) -> str:
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
    try:
        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        raise RuntimeError(f"Couldn't reach that page ({e}).")
    return response.text


def get_usable_tables(html: str) -> list:
    try:
        raw_tables = pd.read_html(StringIO(html))
    except ValueError:
        return []

    usable = []
    for t in raw_tables:
        t_clean = t.dropna(how="all")
        has_named_cols = not all(isinstance(c, int) for c in t_clean.columns)
        is_navbox = any(str(c).strip().lower().startswith("vte") for c in t_clean.columns)
        is_authority_control = any("authority control" in str(c).lower() for c in t_clean.columns)
        if (has_named_cols and not is_navbox and not is_authority_control
                and len(t_clean) >= 3 and len(t_clean.columns) >= 2):
            t_clean = t_clean.copy()
            t_clean.columns = [str(c) for c in t_clean.columns]
            usable.append(t_clean)
    return usable


def _normalize_text(text: str) -> str:
    text = str(text)
    text = re.sub(r"[\x00-\x1f\x7f]+", " ", text)
    text = re.sub(r"\[.*?\]", "", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _flatten_column_name(col) -> str:
    if isinstance(col, tuple):
        parts = [str(p) for p in col if str(p) and "Unnamed" not in str(p)]
        return " ".join(parts) if parts else "Column"
    text = _normalize_text(col)
    return text or "Column"


def _find_matching_column(table: pd.DataFrame, target: str):
    if target in table.columns:
        return target

    normalized_target = _normalize_text(target).lower()
    if not normalized_target:
        return None

    candidates = []
    for col in table.columns:
        normalized_col = _normalize_text(col).lower()
        if normalized_col == normalized_target:
            return col
        if normalized_target in normalized_col or normalized_col in normalized_target:
            candidates.append((len(normalized_col), col))

    if candidates:
        # Prefer the shortest matching normalized column name
        return sorted(candidates, key=lambda item: item[0])[0][1]

    return None


def _describe_table(table: pd.DataFrame, index: int) -> str:
    cols = [_flatten_column_name(c) for c in table.columns[:4]]
    label_guess = cols[0] if cols else "items"
    value_guess = cols[1] if len(cols) > 1 else "values"
    extra = f" + {len(table.columns) - 2} more" if len(table.columns) > 2 else ""
    return f"Table {index + 1}: {label_guess} / {value_guess}{extra}\n({len(table)} rows)"


def choose_table(tables: list, advanced: bool) -> pd.DataFrame:
    if len(tables) == 1:
        info("Found 1 usable table on this page - using it.")
        return tables[0]

    descriptions = [_describe_table(table, i) for i, table in enumerate(tables)]

    idx = popup_ask_from_list(
        "Choose a table",
        f"This page has {len(tables)} usable tables. Pick the one you want to chart:",
        descriptions,
    )
    return tables[idx]


def analyze_table(client, table: pd.DataFrame) -> ChartAnalysis:
    prompt = f"""
This table has these columns: {list(table.columns)}
Sample rows:
{table.head(5).to_string()}
Total rows in table: {len(table)}

You are a senior data visualization designer. Analyze this table using the following meta-prompting steps:

STEP 1: Schema Assessment
- Map the categorical index column ('label_column').
- Identify numeric tracks. If columns track sequential points in time (e.g., years, quarters, time sequences), include ALL of them in 'value_columns' to enable multi-series comparison.

STEP 2: Chart Type Selection
- Rule A (Line): If 'value_columns' contains multiple columns tracking a timeline sequence, or if the 'label_column' represents years/dates, you MUST select 'line'. Do not default to bar.
- Rule B (Pie): If there is exactly 1 value column representing percentages/fractions of a whole and total items <= 8, select 'pie'.
- Rule C (Bar): For discrete entity comparisons, rankings, or flat categories that do not change over a continuous timeline, select 'bar'.

STEP 3: Labeling and Context Engineering
- Provide clear, specific axis titles ('x_axis_title', 'y_axis_title'). Never leave them blank or generic.
- Determine if a legend is required ('show_legend'). Multi-series data tracking multiple lines or grouped columns must always set this to True.
"""
    response = client.models.generate_content(
        model=MODEL_NAME,
        contents=prompt,
        config={"response_mime_type": "application/json", "response_schema": ChartAnalysis}
    )
    return ChartAnalysis.model_validate_json(response.text)


def clean_table_data(table: pd.DataFrame, label_col: str, value_cols: list, title: str) -> dict:
    label_match = _find_matching_column(table, label_col)
    if label_match is None:
        available_columns = [_flatten_column_name(c) for c in table.columns]
        warn(
            "Label column not found. "
            f"Tried '{label_col}', available columns: {available_columns}. "
            "Falling back to the first column as the label column."
        )
        label_match = table.columns[0]

    resolved_value_cols = []
    for col_name in value_cols:
        match = _find_matching_column(table, col_name)
        if match is not None:
            resolved_value_cols.append(match)

    if not resolved_value_cols:
        raise KeyError(f"No numeric value columns matched the requested columns: {value_cols}")

    df = table[[label_match] + resolved_value_cols].copy()
    df.columns = [_flatten_column_name(c) for c in df.columns]

    label_column_name = df.columns[0]
    value_column_names = df.columns[1:]

    for col in value_column_names:
        df[col] = pd.to_numeric(
            df[col].astype(str).str.replace(r"[^\d.\-]", "", regex=True), errors="coerce"
        )
    df = df.dropna(subset=value_column_names)

    return {
        "labels": df[label_column_name].astype(str).tolist(),
        "series": {col: df[col].tolist() for col in value_column_names},
        "title": title
    }


def extract_page_text(html: str) -> str:
    soup = BeautifulSoup(html, "html.parser")
    return soup.get_text(separator=" ", strip=True)[:8000]


def analyze_page_text(client, page_text: str) -> FlatDataAnalysis:
    prompt = f"""
Read this text directly and find any structured metrics or rankings to extract. Follow structural design guidelines for labels.
Webpage text:
{page_text}
"""
    response = client.models.generate_content(
        model=MODEL_NAME,
        contents=prompt,
        config={"response_mime_type": "application/json", "response_schema": FlatDataAnalysis}
    )
    return FlatDataAnalysis.model_validate_json(response.text)


def run_webpage_flow(client, url: str, advanced: bool):
    step("Fetching the page...")
    html = fetch_html(url)
    tables = get_usable_tables(html)

    if tables:
        chosen_table = choose_table(tables, advanced)
        if advanced:
            step("Preview of chosen table:")
            print(chosen_table.head())

        step("Analyzing columns with Gemini (Meta-Prompting)...")
        analysis = analyze_table(client, chosen_table)
        if advanced:
            info(f"Selected: labels = '{analysis.label_column}', values = {analysis.value_columns}")

        data = clean_table_data(chosen_table, analysis.label_column, analysis.value_columns, analysis.title)
        chart_type = choose_chart_type(analysis.recommended, analysis.reason)
    else:
        step("No structural data tables discovered. Falling back to textual parsing extraction...")
        page_text = extract_page_text(html)
        raw_analysis = analyze_page_text(client, page_text)

        # Format the flat analysis to carry structural tracking data attributes
        analysis = ChartAnalysis(
            label_column="Category",
            value_columns=["Values"],
            title=raw_analysis.title,
            x_axis_title=raw_analysis.x_axis_title,
            y_axis_title=raw_analysis.y_axis_title,
            recommended=raw_analysis.recommended,
            show_legend=raw_analysis.show_legend,
            reason=raw_analysis.reason
        )
        data = {
            "labels": [str(l) for l in raw_analysis.labels],
            "series": {"Values": [float(v) for v in raw_analysis.values]},
            "title": raw_analysis.title,
        }
        chart_type = choose_chart_type(analysis.recommended, analysis.reason)

    if not data["labels"]:
        warn("Couldn't extract formatting structures out of this context.")
        return None, None

    count = choose_item_count(data, chart_type)
    return trim_data(data, count), (analysis, chart_type)


# ============================================
# YOUTUBE TRACKING WORKFLOW
# ============================================
def download_youtube_video(url: str) -> str:
    if not YT_DLP_AVAILABLE:
        raise RuntimeError("yt-dlp missing. Execute: python -m pip install yt-dlp")

    step("Downloading video asset footprint (Processing)...")
    ydl_opts = {
        "format": "mp4[height<=480]/mp4",
        "outtmpl": "temp_video.%(ext)s",
        "quiet": True,
    }
    try:
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            info_dict = ydl.extract_info(url, download=True)
            filename = ydl.prepare_filename(info_dict)
    except Exception as e:
        raise RuntimeError(f"Download thread runtime crash context: ({e}).")
    return filename


def analyze_video(client, video_path: str) -> FlatDataAnalysis:
    step("Uploading tracking video block...")
    video_file = client.files.upload(file=video_path)

    while video_file.state.name == "PROCESSING":
        time.sleep(3)
        video_file = client.files.get(name=video_file.name)

    if video_file.state.name == "FAILED":
        raise RuntimeError("Cloud asset mapping failed compilation on service endpoint backend.")

    step("Executing target validation analysis processing logs...")
    prompt = "Review visual parameters carefully. Export any clean statistical rankings data arrays found."

    response = client.models.generate_content(
        model=MODEL_NAME,
        contents=[video_file, prompt],
        config={"response_mime_type": "application/json", "response_schema": FlatDataAnalysis}
    )
    analysis = FlatDataAnalysis.model_validate_json(response.text)

    try:
        client.files.delete(name=video_file.name)
        os.remove(video_path)
    except Exception:
        pass

    return analysis


def run_youtube_flow(client, url: str, advanced: bool):
    video_path = download_youtube_video(url)
    raw_analysis = analyze_video(client, video_path)

    analysis = ChartAnalysis(
        label_column="Category",
        value_columns=["Values"],
        title=raw_analysis.title,
        x_axis_title=raw_analysis.x_axis_title,
        y_axis_title=raw_analysis.y_axis_title,
        recommended=raw_analysis.recommended,
        show_legend=raw_analysis.show_legend,
        reason=raw_analysis.reason
    )

    data = {
        "labels": [str(l) for l in raw_analysis.labels],
        "series": {"Values": [float(v) for v in raw_analysis.values]},
        "title": raw_analysis.title,
    }

    if not data["labels"]:
        warn("No clean tracking data metrics discovered inside target execution space loop context.")
        return None, None

    chart_type = choose_chart_type(analysis.recommended, analysis.reason)
    count = choose_item_count(data, chart_type)
    return trim_data(data, count), (analysis, chart_type)


# ============================================
# INTERFACE MAIN LOOP ENTRY POINT
# ============================================
def main():
    client = genai.Client(api_key=API_KEY)

    print("=" * 50)
    print("   Advanced Meta-Prompting Chart Engine Generator")
    print("=" * 50)

    url = popup_ask_text("Link -> Chart Tool", "Paste a webpage or YouTube link:").strip()
    if not url:
        return

    mode = popup_ask_choice("View mode", "Choose a view mode:", ["simple", "advanced"])
    advanced = mode == "advanced"

    try:
        if is_youtube_url(url):
            result, meta = run_youtube_flow(client, url, advanced)
        else:
            result, meta = run_webpage_flow(client, url, advanced)
    except RuntimeError as e:
        popup_warn("Error", str(e))
        return

    if result is None:
        return

    analysis, chart_type = meta
    step("Generating visualization output components...")
    save_chart_image(result, analysis, chart_type)
    save_excel(result, analysis, chart_type)
    popup_info("Done", "Your chart image and Excel file were created in the project folder.")


if __name__ == "__main__":
    main()